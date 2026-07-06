"""
Persistent travel time cache backed by an Excel file.

Stores Google Maps API results with contextual fields so they can be
reused for the same day-of-week and time-of-day combinations without
re-calling the API.

Cache key logic:
- Same origin/destination (rounded to 4dp ~11m)
- Same day of week (traffic patterns repeat weekly)
- Same hour of day (traffic varies by time)

File: Data/travel_time_cache.xlsx
"""

import os
import math
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional

log = logging.getLogger(__name__)

CACHE_FILE = Path(__file__).resolve().parent.parent.parent / "Data" / "travel_time_cache.xlsx"

# In-memory cache loaded from Excel at startup
_persistent_cache: dict[tuple, dict] = {}
_cache_dirty = False  # Track if new entries need to be saved


def _make_key(lat1: float, lon1: float, lat2: float, lon2: float,
              day_of_week: int, hour_of_day: int) -> tuple:
    """Create cache key: coords rounded to 4dp + day of week + hour bucket."""
    return (
        round(lat1, 4), round(lon1, 4),
        round(lat2, 4), round(lon2, 4),
        day_of_week,    # 0=Monday, 6=Sunday
        hour_of_day,    # 0-23
    )


def load_cache():
    """Load the persistent cache from Excel file into memory."""
    global _persistent_cache
    _persistent_cache = {}

    if not CACHE_FILE.exists():
        log.info("No persistent travel cache file found — starting fresh")
        return

    try:
        import openpyxl
        wb = openpyxl.load_workbook(CACHE_FILE, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

        for row in rows:
            if len(row) < 12:
                continue
            (origin_lat, origin_lon, dest_lat, dest_lon,
             day_of_week, hour_of_day, travel_mins, distance_miles,
             haversine_miles_val, source, created_at, terminal_name) = row[:12]

            if travel_mins is None:
                continue

            key = _make_key(
                float(origin_lat), float(origin_lon),
                float(dest_lat), float(dest_lon),
                int(day_of_week), int(hour_of_day),
            )
            _persistent_cache[key] = {
                "travel_mins": float(travel_mins),
                "distance_miles": float(distance_miles) if distance_miles else 0,
                "source": source or "google_maps",
            }

        wb.close()
        log.info(f"Loaded {len(_persistent_cache)} entries from travel cache")
    except Exception as e:
        log.warning(f"Could not load travel cache: {e}")


def save_cache():
    """Save the in-memory cache to Excel file."""
    global _cache_dirty
    if not _cache_dirty:
        return

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Travel Time Cache"

        # Header
        ws.append([
            "origin_lat", "origin_lon", "dest_lat", "dest_lon",
            "day_of_week", "hour_of_day", "travel_mins", "distance_miles",
            "haversine_miles", "source", "created_at", "terminal_name",
        ])

        for key, data in _persistent_cache.items():
            lat1, lon1, lat2, lon2, dow, hour = key
            haversine = _haversine_miles(lat1, lon1, lat2, lon2)
            ws.append([
                lat1, lon1, lat2, lon2,
                dow, hour,
                round(data["travel_mins"], 2),
                round(data.get("distance_miles", 0), 2),
                round(haversine, 2),
                data.get("source", "google_maps"),
                data.get("created_at", datetime.now().isoformat()),
                data.get("terminal_name", ""),
            ])

        # Auto-size columns
        CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        wb.save(CACHE_FILE)
        _cache_dirty = False
        log.info(f"Saved {len(_persistent_cache)} entries to {CACHE_FILE}")
    except Exception as e:
        log.warning(f"Could not save travel cache: {e}")


def _haversine_miles(lat1, lon1, lat2, lon2):
    R = 3958.8
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
    return 2 * R * math.asin(math.sqrt(a))


def get_cached_travel_mins(
    lat1: float, lon1: float,
    lat2: float, lon2: float,
    departure_epoch: Optional[int] = None,
) -> Optional[float]:
    """
    Look up travel time from persistent cache.
    Returns travel_mins if found, None if cache miss.
    """
    if not _persistent_cache:
        return None

    if departure_epoch:
        dt = datetime.fromtimestamp(departure_epoch)
        dow = dt.weekday()
        hour = dt.hour
    else:
        # Use current time if no departure specified
        now = datetime.now()
        dow = now.weekday()
        hour = now.hour

    key = _make_key(lat1, lon1, lat2, lon2, dow, hour)
    entry = _persistent_cache.get(key)
    if entry:
        return entry["travel_mins"]
    return None


def store_travel_time(
    lat1: float, lon1: float,
    lat2: float, lon2: float,
    departure_epoch: Optional[int],
    travel_mins: float,
    distance_miles: float = 0,
    source: str = "google_maps",
    terminal_name: str = "",
):
    """
    Store a travel time result in the persistent cache.
    """
    global _cache_dirty

    if departure_epoch:
        dt = datetime.fromtimestamp(departure_epoch)
        dow = dt.weekday()
        hour = dt.hour
    else:
        now = datetime.now()
        dow = now.weekday()
        hour = now.hour

    key = _make_key(lat1, lon1, lat2, lon2, dow, hour)
    _persistent_cache[key] = {
        "travel_mins": travel_mins,
        "distance_miles": distance_miles,
        "source": source,
        "created_at": datetime.now().isoformat(),
        "terminal_name": terminal_name,
    }
    _cache_dirty = True


def cache_size() -> int:
    """Return current number of cached entries."""
    return len(_persistent_cache)


def is_loaded() -> bool:
    """Check if cache has been loaded."""
    return len(_persistent_cache) > 0 or CACHE_FILE.exists()
